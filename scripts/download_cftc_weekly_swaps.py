#!/usr/bin/env python3
"""Download archived CFTC Weekly Swaps Report FX tables.

The official archive stores one XLSX workbook per report edition.  FX tables
are present only from the 2018-10-24 workbook onward.  These aggregates are
not spot-FX order flow, signed positioning, prices, forward points, or carry.

The downloader intentionally keeps every normalized row ineligible for strict
PIT promotion.  The archive calls its links previous publications, but the
files were not cryptographically captured by this project at original release
and the archive date is demonstrably not the actual publication date for the
2025 appropriations catch-up editions.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import time
import warnings
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from datetime import time as datetime_time
from html.parser import HTMLParser
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from urllib.parse import unquote, urljoin, urlparse
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook

PROGRAM_VERSION = "cftc-weekly-swaps-fx-v1"
HOST = "www.cftc.gov"
BASE_URL = f"https://{HOST}"
ARCHIVE_URL = f"{BASE_URL}/MarketReports/SwapsReports/Archive/index.htm"
RELEASE_SCHEDULE_URL = f"{BASE_URL}/MarketReports/SwapsReports/ReleaseSchedule/index.htm"
EXPLANATORY_NOTES_URL = f"{BASE_URL}/MarketReports/SwapsReports/ExplanatoryNotes/index.htm"
WEB_POLICY_URL = f"{BASE_URL}/webpolicy/index.htm"

FIRST_FX_WORKBOOK = date(2018, 10, 24)
LAST_RESEARCH_EDITION = date(2025, 12, 29)
KNOWN_NO_REPORT_LINKS = frozenset({date(2018, 12, 26)})
EXPECTED_REPORT_COUNTS = {
    2018: 9,
    2019: 47,
    2020: 52,
    2021: 52,
    2022: 52,
    2023: 51,
    2024: 21,
    2025: 52,
}
EXPECTED_ROW_LABELS = (
    "USD/",
    "EUROPE",
    "EUR",
    "GBP",
    "CHF",
    "OTHER2",
    "AMERICAS/CARIBBEAN",
    "CAD",
    "BRL",
    "OTHER3",
    "ASIA/PACIFIC",
    "JPY",
    "CNY",
    "KRW",
    "HKD",
    "AUD",
    "OTHER4",
    "OTHER5",
    "EUR/non-USD",
    "OTHER6",
    "TOTAL",
)
DISCLOSED_CURRENCIES = frozenset(
    {"EUR", "GBP", "CHF", "CAD", "BRL", "JPY", "CNY", "KRW", "HKD", "AUD"}
)
CORE_RESEARCH_CURRENCIES = frozenset({"EUR", "GBP", "CHF", "CAD", "JPY", "AUD"})
MISSING_CORE_CURRENCIES = ("NZD",)

OLD_PRODUCT_HEADERS = (
    "Currency Pair1",
    "Swaps & Forwards",
    "NDF7",
    "Options",
    "Exotics",
    "Cross Currency",
    "Total",
)
OLD_PRODUCT_HEADERS_SINGULAR_EXOTIC = (
    "Currency Pair1",
    "Swaps & Forwards",
    "NDF7",
    "Options",
    "EXOTIC",
    "Cross Currency",
    "Total",
)
NEW_PRODUCT_HEADERS = (
    "Currency Pair1",
    "Swaps & Forwards",
    "NDF7",
    "Options",
    "Other8",
    "Total",
)
PRODUCT_NAMES = {
    "Swaps & Forwards": "swaps_and_forwards",
    "NDF7": "ndf",
    "Options": "options",
    "Exotics": "exotics",
    "EXOTIC": "exotics",
    "Cross Currency": "cross_currency",
    "Other8": "other_includes_cross_currency_and_exotics",
    "Total": "total",
}
TABLES = {
    "19b": ("gross_notional_outstanding", "usd_millions"),
    "20b": ("transaction_ticket_volume", "ticket_count"),
    "21b": ("transaction_dollar_volume", "usd_millions"),
}
OUTPUT_COLUMNS = (
    "edition_date",
    "reporting_period_end",
    "availability_time",
    "availability_quality",
    "release_date_quality",
    "release_time_quality",
    "strict_pit_eligible",
    "value_vintage_quality",
    "methodology_regime",
    "product_schema_regime",
    "table_id",
    "metric",
    "currency",
    "currency_pair_bucket",
    "product_category",
    "value",
    "unit",
    "source_url",
    "source_sha256",
)
MAX_HTML_BYTES = 2 * 1024 * 1024
MAX_XLSX_BYTES = 2 * 1024 * 1024
NEW_YORK = ZoneInfo("America/New_York")
_DATE_IN_FILENAME = re.compile(r"(?i)cftc_swaps_report_(\d{2})_(\d{2})_(\d{4})(?:[^0-9]|$)")
_SHA256 = re.compile(r"[0-9a-f]{64}")


@dataclass(frozen=True)
class ReportLink:
    edition_date: date
    url: str

    @property
    def source_id(self) -> str:
        return f"weekly_swaps_{self.edition_date:%Y%m%d}"


class _ArchiveParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[str] = []
        self.text: list[str] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag.lower() != "a":
            return
        href = dict(attrs).get("href") or ""
        if ".xlsx" in href.lower():
            self.links.append(href)

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.text.append(data.strip())


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _iso(value: datetime) -> str:
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")


def _safe_url(raw: str, *, expected_suffix: str | None = None) -> str:
    url = urljoin(BASE_URL, raw)
    parsed = urlparse(url)
    decoded_path = unquote(parsed.path)
    if parsed.scheme != "https" or parsed.hostname != HOST or parsed.query or parsed.fragment:
        raise ValueError(f"refusing non-canonical CFTC URL: {url}")
    allowed = decoded_path.startswith("/sites/default/files/") or decoded_path.startswith(
        "/idc/groups/public/@swapsreport/documents/file/"
    )
    if expected_suffix and not decoded_path.lower().endswith(expected_suffix):
        raise ValueError(f"CFTC source has unexpected suffix: {url}")
    if expected_suffix == ".xlsx" and not allowed:
        raise ValueError(f"CFTC workbook escaped the official swaps-report archive: {url}")
    return url


def discover_reports(
    payload: bytes,
    *,
    expected_counts: Mapping[int, int] | None = EXPECTED_REPORT_COUNTS,
) -> tuple[ReportLink, ...]:
    try:
        text = payload.decode("utf-8")
    except UnicodeDecodeError as error:
        raise ValueError("CFTC Weekly Swaps archive is not valid UTF-8") from error
    parser = _ArchiveParser()
    parser.feed(text)
    evidence_text = " ".join(parser.text)
    required_phrases = (
        "previous publications",
        "On October 17, 2018 FX swaps were first included",
        "did not issue a Weekly Swaps Report during this period",
    )
    if not all(phrase in evidence_text for phrase in required_phrases):
        raise ValueError("CFTC Weekly Swaps archive evidence wording drifted")

    discovered: dict[date, str] = {}
    for href in parser.links:
        url = _safe_url(href, expected_suffix=".xlsx")
        filename = unquote(urlparse(url).path).rsplit("/", 1)[-1]
        match = _DATE_IN_FILENAME.search(filename)
        if match is None:
            continue
        month, day, year = (int(token) for token in match.groups())
        try:
            edition = date(year, month, day)
        except ValueError as error:
            raise ValueError(f"invalid CFTC workbook date in {url}") from error
        if edition < FIRST_FX_WORKBOOK or edition > LAST_RESEARCH_EDITION:
            continue
        if edition in KNOWN_NO_REPORT_LINKS:
            continue
        prior = discovered.get(edition)
        if prior is not None and prior != url:
            raise ValueError(f"duplicate CFTC Weekly Swaps URLs for {edition}: {prior}, {url}")
        discovered[edition] = url

    reports = tuple(ReportLink(day, discovered[day]) for day in sorted(discovered))
    if not reports:
        raise ValueError("CFTC Weekly Swaps archive exposed no FX-era workbooks")
    if expected_counts is not None:
        actual = {
            year: sum(report.edition_date.year == year for report in reports)
            for year in expected_counts
        }
        if actual != dict(expected_counts):
            raise ValueError(
                "CFTC Weekly Swaps archive coverage drifted; "
                f"expected={dict(expected_counts)}, actual={actual}"
            )
    return reports


def _workbook_dates(workbook: Any) -> tuple[date, date]:
    sheet = workbook["Table of Contents"]
    found: dict[str, date] = {}
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Unknown extension is not supported and will be removed",
            category=UserWarning,
        )
        for row in sheet.iter_rows(values_only=True):
            for index, value in enumerate(row):
                if value not in {"Release Date:", "Reporting Period As Of:"}:
                    continue
                candidates = [cell for cell in row[index + 1 :] if cell not in {None, ""}]
                if len(candidates) != 1:
                    raise ValueError(f"workbook {value} does not have one adjacent value")
                raw = candidates[0]
                if isinstance(raw, datetime):
                    parsed = raw.date()
                elif isinstance(raw, date):
                    parsed = raw
                else:
                    try:
                        parsed = date.fromisoformat(str(raw).strip())
                    except ValueError as error:
                        raise ValueError(f"workbook {value} is not a date") from error
                found[str(value)] = parsed
    if set(found) != {"Release Date:", "Reporting Period As Of:"}:
        raise ValueError("workbook is missing its edition/reporting-period dates")
    return found["Release Date:"], found["Reporting Period As Of:"]


def _schema(headers: tuple[str, ...]) -> str:
    if headers in {OLD_PRODUCT_HEADERS, OLD_PRODUCT_HEADERS_SINGULAR_EXOTIC}:
        return "separate_exotics_cross_currency"
    if headers == NEW_PRODUCT_HEADERS:
        return "merged_other_includes_cross_currency_and_exotics"
    raise ValueError(f"unknown CFTC Weekly Swaps FX currency-table schema: {headers}")


def _methodology_regime(edition: date) -> str:
    if edition < date(2022, 12, 12):
        return "fx_launch_pre_2022_12_09_methodology_improvement"
    return "post_2022_12_09_methodology_improvement"


def _integer(value: object, *, field: str) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{field} is not numeric")
    parsed = int(value)
    if float(value) != parsed or parsed < 0:
        raise ValueError(f"{field} must be a non-negative integer")
    return parsed


def _availability_fields(edition: date, retrieved_at: datetime) -> dict[str, object]:
    catch_up = date(2025, 10, 6) <= edition <= date(2025, 11, 10)
    if catch_up:
        return {
            "availability_time": _iso(retrieved_at),
            "availability_quality": (
                "first_project_retrieval_after_official_appropriations_catch_up_"
                "actual_publication_mapping_unavailable"
            ),
            "release_date_quality": "edition_label_not_actual_catch_up_publication_date",
            "release_time_quality": "actual_publication_time_unverified",
        }
    next_local_day = datetime.combine(
        edition + timedelta(days=1), datetime_time.min, tzinfo=NEW_YORK
    )
    return {
        "availability_time": _iso(next_local_day),
        "availability_quality": (
            "next_new_york_calendar_day_after_official_archive_publication_date_"
            "not_actual_timestamp"
        ),
        "release_date_quality": "official_archive_publication_date",
        "release_time_quality": "general_release_rule_not_actual_timestamp",
    }


def parse_workbook(
    report: ReportLink,
    payload: bytes,
    *,
    retrieved_at: datetime,
) -> list[dict[str, object]]:
    if len(payload) > MAX_XLSX_BYTES:
        raise ValueError(f"{report.source_id}: XLSX exceeds size limit")
    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            if archive.testzip() is not None or "xl/workbook.xml" not in archive.namelist():
                raise ValueError(f"{report.source_id}: XLSX ZIP integrity failed")
    except BadZipFile as error:
        raise ValueError(f"{report.source_id}: response is not an XLSX ZIP") from error
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
                category=UserWarning,
            )
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except (OSError, ValueError, BadZipFile) as error:
        raise ValueError(f"{report.source_id}: cannot parse XLSX") from error
    required_sheets = {"Table of Contents", *TABLES}
    if not required_sheets.issubset(workbook.sheetnames):
        missing = sorted(required_sheets - set(workbook.sheetnames))
        raise ValueError(f"{report.source_id}: workbook lacks sheets {missing}")
    edition, reporting_period = _workbook_dates(workbook)
    if edition != report.edition_date:
        raise ValueError(
            f"{report.source_id}: URL/archive date {report.edition_date} != workbook {edition}"
        )
    lag = (edition - reporting_period).days
    if reporting_period.weekday() != 4 or not 10 <= lag <= 25:
        raise ValueError(f"{report.source_id}: implausible reporting-period date/lag")

    digest = _sha256(payload)
    availability = _availability_fields(edition, retrieved_at)
    rows: list[dict[str, object]] = []
    schemas: set[str] = set()
    methodology = _methodology_regime(edition)
    for table_id, (metric, unit) in TABLES.items():
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
                category=UserWarning,
            )
            raw_rows = [
                tuple(value for value in row)
                for row in workbook[table_id].iter_rows(values_only=True)
                if any(value not in {None, ""} for value in row)
            ]
        if len(raw_rows) != 23:
            raise ValueError(f"{report.source_id}/{table_id}: expected header, 21 rows and notes")
        headers = tuple(str(value).strip() for value in raw_rows[0])
        schema_kind = _schema(headers)
        schemas.add(schema_kind)
        data_rows = raw_rows[1:22]
        labels = tuple(str(row[0]).strip() for row in data_rows)
        if labels != EXPECTED_ROW_LABELS:
            raise ValueError(f"{report.source_id}/{table_id}: currency row labels drifted")
        product_headers = headers[1:]
        for raw_row in data_rows:
            label = str(raw_row[0]).strip()
            values = [
                _integer(value, field=f"{report.source_id}/{table_id}/{label}/{header}")
                for header, value in zip(product_headers, raw_row[1:], strict=True)
            ]
            # Published USD-million and ticket cells are independently rounded;
            # audited official workbooks can differ from their displayed Total by 1-2.
            if abs(sum(values[:-1]) - values[-1]) > 2:
                raise ValueError(f"{report.source_id}/{table_id}/{label}: Total does not add up")
            if label not in DISCLOSED_CURRENCIES:
                continue
            for product_header, value in zip(product_headers, values, strict=True):
                rows.append(
                    {
                        "edition_date": edition.isoformat(),
                        "reporting_period_end": reporting_period.isoformat(),
                        **availability,
                        "strict_pit_eligible": False,
                        "value_vintage_quality": (
                            "official_archived_edition_not_cryptographically_verified_at_"
                            "original_publication"
                        ),
                        "methodology_regime": methodology,
                        "product_schema_regime": schema_kind,
                        "table_id": table_id,
                        "metric": metric,
                        "currency": label,
                        "currency_pair_bucket": f"USD/{label}_aggregate_no_direction",
                        "product_category": PRODUCT_NAMES[product_header],
                        "value": value,
                        "unit": unit,
                        "source_url": report.url,
                        "source_sha256": digest,
                    }
                )
    if len(schemas) != 1:
        raise ValueError(f"{report.source_id}: FX tables disagree on their schema regime")
    expected_products = 6 if next(iter(schemas)) == "separate_exotics_cross_currency" else 5
    expected_rows = len(TABLES) * len(DISCLOSED_CURRENCIES) * expected_products
    if len(rows) != expected_rows:
        raise ValueError(f"{report.source_id}: normalized row count is inconsistent")
    return rows


def _atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(dir=path.parent, delete=False) as handle:
        handle.write(payload)
        handle.flush()
        os.fsync(handle.fileno())
        temporary = Path(handle.name)
    temporary.replace(path)


def _manifest_path(root: Path, value: object, *, field: str) -> Path:
    if not isinstance(value, str) or not value or Path(value).is_absolute():
        raise ValueError(f"{field} is not a safe relative path")
    path = (root / value).resolve()
    if not path.is_relative_to(root.resolve()):
        raise ValueError(f"{field} escapes the output directory")
    return path


def _load_prior_sources(root: Path) -> dict[str, dict[str, object]]:
    manifest_path = root / "cftc_weekly_swaps_manifest.json"
    sidecar = manifest_path.with_suffix(".sha256")
    cache_root = root / "raw" / "cache"
    if not manifest_path.exists() and not sidecar.exists():
        if cache_root.exists() and any(path.is_file() for path in cache_root.rglob("*")):
            raise ValueError("orphan CFTC Weekly Swaps cache exists without a prior manifest")
        return {}
    if (
        not manifest_path.is_file()
        or not sidecar.is_file()
        or manifest_path.is_symlink()
        or sidecar.is_symlink()
    ):
        raise ValueError("prior Weekly Swaps manifest and hash must be regular files")
    manifest_payload = manifest_path.read_bytes()
    if sidecar.read_text(encoding="ascii").strip().split() != [_sha256(manifest_payload)]:
        raise ValueError("prior Weekly Swaps manifest SHA-256 verification failed")
    try:
        manifest = json.loads(manifest_payload)
    except (UnicodeError, json.JSONDecodeError) as error:
        raise ValueError("prior Weekly Swaps manifest is unreadable") from error
    if (
        not isinstance(manifest, dict)
        or manifest.get("schema_version") != 1
        or manifest.get("program_version") != PROGRAM_VERSION
        or manifest.get("dataset_kind")
        != "cftc_weekly_swaps_fx_activity_not_spot_order_flow_not_directional"
    ):
        raise ValueError("prior Weekly Swaps manifest contract is invalid")
    normalized = _manifest_path(root, manifest.get("normalized_path"), field="normalized_path")
    digest = manifest.get("normalized_sha256")
    if (
        not isinstance(digest, str)
        or _SHA256.fullmatch(digest) is None
        or normalized.is_symlink()
        or not normalized.is_file()
        or _sha256(normalized.read_bytes()) != digest
    ):
        raise ValueError("prior normalized Weekly Swaps output hash verification failed")
    raw_sources = manifest.get("sources")
    if not isinstance(raw_sources, list) or not raw_sources:
        raise ValueError("prior Weekly Swaps manifest has no sources")
    sources: dict[str, dict[str, object]] = {}
    expected_cache: set[Path] = set()
    for index, raw in enumerate(raw_sources):
        if not isinstance(raw, dict):
            raise ValueError(f"prior source {index} is invalid")
        source_id = raw.get("source_id")
        if not isinstance(source_id, str) or not source_id or source_id in sources:
            raise ValueError(f"prior source_id {index} is invalid")
        for field in ("cache_path", "cache_metadata_path"):
            expected_cache.add(_manifest_path(root, raw.get(field), field=field))
        sources[source_id] = raw
    if cache_root.exists():
        actual_cache = {path.resolve() for path in cache_root.rglob("*") if path.is_file()}
        if actual_cache != expected_cache:
            raise ValueError("Weekly Swaps cache contains orphan or missing files")
    return sources


def _fetch(
    url: str,
    *,
    timeout: float,
    retries: int,
    maximum_bytes: int,
    transport: httpx.BaseTransport | None,
    sleep: Callable[[float], None],
) -> tuple[bytes, dict[str, str]]:
    parsed = urlparse(url)
    if parsed.scheme != "https" or parsed.hostname != HOST:
        raise ValueError(f"refusing non-CFTC URL: {url}")
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with httpx.Client(
                transport=transport,
                follow_redirects=True,
                timeout=timeout,
                headers={"User-Agent": f"{PROGRAM_VERSION} (+public research archive)"},
            ) as client:
                with client.stream("GET", url) as response:
                    response.raise_for_status()
                    if response.url.scheme != "https" or response.url.host != HOST:
                        raise ValueError("CFTC Weekly Swaps redirect escaped the official host")
                    declared = response.headers.get("Content-Length")
                    if declared and int(declared) > maximum_bytes:
                        raise ValueError("CFTC Weekly Swaps response exceeds declared size limit")
                    chunks: list[bytes] = []
                    size = 0
                    for chunk in response.iter_bytes():
                        size += len(chunk)
                        if size > maximum_bytes:
                            raise ValueError("CFTC Weekly Swaps response exceeds size limit")
                        chunks.append(chunk)
                    payload = b"".join(chunks)
                    if not payload:
                        raise ValueError("CFTC Weekly Swaps response is empty")
                    return payload, {
                        "content_type": response.headers.get("Content-Type", ""),
                        "etag": response.headers.get("ETag", ""),
                        "last_modified": response.headers.get("Last-Modified", ""),
                    }
        except ValueError:
            raise
        except httpx.HTTPStatusError as error:
            last_error = error
            if error.response.status_code < 500 and error.response.status_code != 429:
                raise
            if attempt < retries:
                retry_after = error.response.headers.get("Retry-After", "").strip()
                backoff = 0.75 * (2**attempt)
                delay = max(float(retry_after), backoff) if retry_after.isdigit() else backoff
                sleep(min(30.0, delay))
        except (httpx.HTTPError, OSError) as error:
            last_error = error
            if attempt < retries:
                sleep(min(8.0, 0.75 * (2**attempt)))
    assert last_error is not None
    raise last_error


def _load_source(
    root: Path,
    *,
    source_id: str,
    url: str,
    suffix: str,
    maximum_bytes: int,
    prior: dict[str, object] | None,
    refresh: bool,
    timeout: float,
    retries: int,
    transport: httpx.BaseTransport | None,
    sleep: Callable[[float], None],
    retrieved_at: datetime,
) -> tuple[bytes, dict[str, object]]:
    cache = root / "raw" / "cache" / f"{source_id}{suffix}"
    metadata_path = cache.with_suffix(f"{cache.suffix}.meta.json")
    if cache.is_file() and not refresh:
        if prior is None:
            raise ValueError(f"{source_id}: orphan cache lacks a prior manifest record")
        if cache.is_symlink() or metadata_path.is_symlink():
            raise ValueError(f"{source_id}: cache or metadata cannot be a symlink")
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise ValueError(f"{source_id}: cached metadata is invalid") from error
        payload = cache.read_bytes()
        digest = _sha256(payload)
        identity = {
            "source_id": source_id,
            "url": url,
            "cache_path": cache.relative_to(root).as_posix(),
            "cache_metadata_path": metadata_path.relative_to(root).as_posix(),
        }
        if any(prior.get(field) != value for field, value in identity.items()):
            raise ValueError(f"{source_id}: prior cache identity does not match")
        for field in (
            "source_id",
            "url",
            "retrieved_at",
            "bytes",
            "sha256",
            "content_type",
            "etag",
            "last_modified",
        ):
            if metadata.get(field) != prior.get(field):
                raise ValueError(f"{source_id}: metadata differs from manifest at {field}")
        if metadata.get("bytes") != len(payload) or metadata.get("sha256") != digest:
            raise ValueError(f"{source_id}: cached source hash verification failed")
        source_retrieved = datetime.fromisoformat(
            str(metadata["retrieved_at"]).replace("Z", "+00:00")
        )
        if source_retrieved.tzinfo is None or _iso(source_retrieved) != metadata["retrieved_at"]:
            raise ValueError(f"{source_id}: retrieval timestamp is not canonical UTC")
        archive = _manifest_path(root, prior.get("archive_path"), field="archive_path")
        if archive.is_symlink() or not archive.is_file() or _sha256(archive.read_bytes()) != digest:
            raise ValueError(f"{source_id}: archived snapshot hash mismatch")
        status = "cached"
        response_metadata = {
            field: str(metadata.get(field, ""))
            for field in ("content_type", "etag", "last_modified")
        }
    else:
        if not refresh and (prior is not None or metadata_path.exists()):
            raise ValueError(f"{source_id}: prior cache is incomplete; use --refresh")
        payload, response_metadata = _fetch(
            url,
            timeout=timeout,
            retries=retries,
            maximum_bytes=maximum_bytes,
            transport=transport,
            sleep=sleep,
        )
        source_retrieved = retrieved_at
        status = "downloaded"
        metadata = {
            "source_id": source_id,
            "url": url,
            "retrieved_at": _iso(source_retrieved),
            "bytes": len(payload),
            "sha256": _sha256(payload),
            **response_metadata,
        }
        _atomic_write(cache, payload)
        _atomic_write(metadata_path, (json.dumps(metadata, indent=2) + "\n").encode())
    digest = _sha256(payload)
    archive = (
        root
        / "raw"
        / "archive"
        / source_id
        / f"{source_retrieved:%Y%m%dT%H%M%S%fZ}_{digest[:16]}{suffix}"
    )
    if archive.exists() and _sha256(archive.read_bytes()) != digest:
        raise ValueError(f"{source_id}: archived snapshot hash mismatch")
    if not archive.exists():
        _atomic_write(archive, payload)
    return payload, {
        "source_id": source_id,
        "url": url,
        "status": status,
        "retrieved_at": _iso(source_retrieved),
        "bytes": len(payload),
        "sha256": digest,
        "cache_path": cache.relative_to(root).as_posix(),
        "cache_metadata_path": metadata_path.relative_to(root).as_posix(),
        "archive_path": archive.relative_to(root).as_posix(),
        **response_metadata,
    }


def _write_csv(rows: Sequence[dict[str, object]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=OUTPUT_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        if tuple(row) != OUTPUT_COLUMNS:
            raise ValueError("normalized Weekly Swaps row does not match the frozen schema")
        writer.writerow(row)
    return output.getvalue().encode()


def download(
    output: Path,
    *,
    refresh: bool = False,
    timeout: float = 45.0,
    retries: int = 3,
    delay: float = 0.75,
    transport: httpx.BaseTransport | None = None,
    sleep: Callable[[float], None] = time.sleep,
    retrieved_at: datetime | None = None,
    expected_counts: Mapping[int, int] | None = EXPECTED_REPORT_COUNTS,
) -> dict[str, object]:
    root = output.resolve()
    if output.exists() and output.is_symlink():
        raise ValueError("Weekly Swaps output directory cannot be a symlink")
    root.mkdir(parents=True, exist_ok=True)
    prior = {} if refresh else _load_prior_sources(root)
    now = retrieved_at or datetime.now(UTC)
    if now.tzinfo is None:
        raise ValueError("retrieved_at must include a timezone")
    now = now.astimezone(UTC)
    source_records: list[dict[str, object]] = []

    evidence = (
        ("evidence_archive", ARCHIVE_URL),
        ("evidence_release_schedule", RELEASE_SCHEDULE_URL),
        ("evidence_explanatory_notes", EXPLANATORY_NOTES_URL),
        ("evidence_web_policy", WEB_POLICY_URL),
    )
    evidence_payloads: dict[str, bytes] = {}
    for index, (source_id, url) in enumerate(evidence):
        if index and delay:
            sleep(delay)
        payload, record = _load_source(
            root,
            source_id=source_id,
            url=url,
            suffix=".html",
            maximum_bytes=MAX_HTML_BYTES,
            prior=prior.get(source_id),
            refresh=refresh,
            timeout=timeout,
            retries=retries,
            transport=transport,
            sleep=sleep,
            retrieved_at=now,
        )
        try:
            decoded = payload.decode("utf-8")
        except UnicodeDecodeError as error:
            raise ValueError(f"{source_id}: evidence is not UTF-8") from error
        if "CFTC" not in decoded:
            raise ValueError(f"{source_id}: response is not recognizable CFTC evidence")
        evidence_payloads[source_id] = payload
        source_records.append(record)
    schedule_text = evidence_payloads["evidence_release_schedule"].decode("utf-8")
    if (
        "released at 3:30 p.m. Eastern time on Monday" not in schedule_text
        or "interrupted from October 1" not in schedule_text
        or "resume publication" not in schedule_text
    ):
        raise ValueError("CFTC Weekly Swaps release-schedule evidence wording drifted")
    explanatory_text = evidence_payloads["evidence_explanatory_notes"].decode("utf-8")
    if (
        "Transaction dollar volume represents" not in explanatory_text
        or "Transaction ticket volume represents" not in explanatory_text
        or "Gross notional outstanding represents" not in explanatory_text
    ):
        raise ValueError("CFTC Weekly Swaps explanatory evidence wording drifted")
    policy_text = evidence_payloads["evidence_web_policy"].decode("utf-8")
    if "Government information at the CFTC website is in the public domain" not in policy_text:
        raise ValueError("CFTC public-domain policy evidence wording drifted")

    reports = discover_reports(
        evidence_payloads["evidence_archive"], expected_counts=expected_counts
    )
    normalized: list[dict[str, object]] = []
    for report in reports:
        if delay:
            sleep(delay)
        payload, record = _load_source(
            root,
            source_id=report.source_id,
            url=report.url,
            suffix=".xlsx",
            maximum_bytes=MAX_XLSX_BYTES,
            prior=prior.get(report.source_id),
            refresh=refresh,
            timeout=timeout,
            retries=retries,
            transport=transport,
            sleep=sleep,
            retrieved_at=now,
        )
        rows = parse_workbook(report, payload, retrieved_at=now)
        normalized.extend(rows)
        record["edition_date"] = report.edition_date.isoformat()
        record["reporting_period_end"] = rows[0]["reporting_period_end"]
        record["normalized_rows"] = len(rows)
        source_records.append(record)

    if refresh:
        cache_root = root / "raw" / "cache"
        expected_cache = {
            (root / str(record[field])).resolve()
            for record in source_records
            for field in ("cache_path", "cache_metadata_path")
        }
        for cached_path in cache_root.rglob("*") if cache_root.exists() else ():
            if cached_path.is_symlink():
                raise ValueError("Weekly Swaps refresh refuses symlinks in the cache")
            if cached_path.is_file() and cached_path.resolve() not in expected_cache:
                cached_path.unlink()

    if prior and set(prior) != {str(record["source_id"]) for record in source_records}:
        raise ValueError("prior Weekly Swaps manifest contains stale sources; use --refresh")

    normalized.sort(
        key=lambda row: (
            str(row["edition_date"]),
            str(row["metric"]),
            str(row["currency"]),
            str(row["product_category"]),
        )
    )
    csv_payload = _write_csv(normalized)
    normalized_path = root / "cftc_weekly_swaps_fx_activity.csv"
    _atomic_write(normalized_path, csv_payload)
    counts_by_year = {
        str(year): sum(report.edition_date.year == year for report in reports)
        for year in range(2018, 2026)
    }
    reporting_periods = sorted({str(row["reporting_period_end"]) for row in normalized})
    manifest: dict[str, object] = {
        "schema_version": 1,
        "program_version": PROGRAM_VERSION,
        "dataset_kind": "cftc_weekly_swaps_fx_activity_not_spot_order_flow_not_directional",
        "retrieved_at": _iso(now),
        "edition_start": reports[0].edition_date.isoformat(),
        "edition_end": reports[-1].edition_date.isoformat(),
        "reporting_period_start": reporting_periods[0],
        "reporting_period_end": reporting_periods[-1],
        "report_count": len(reports),
        "report_counts_by_year": counts_by_year,
        "normalized_rows": len(normalized),
        "normalized_path": normalized_path.relative_to(root).as_posix(),
        "normalized_sha256": _sha256(csv_payload),
        "strict_pit_eligible": False,
        "archive_value_quality": (
            "official_archived_editions_not_cryptographically_verified_at_original_release"
        ),
        "release_time_quality": "not_verified_actual_for_all_editions",
        "disclosed_individual_currencies": sorted(DISCLOSED_CURRENCIES),
        "core_research_currencies": sorted(CORE_RESEARCH_CURRENCIES),
        "missing_core_currency_disclosure": list(MISSING_CORE_CURRENCIES),
        "known_archive_issues": [
            "2018-12-26 archive link is excluded because the official page says no report "
            "was issued during the appropriations lapse and the link currently returns 404.",
            "2023-01-30 is not downloaded: the current 2023 archive cell links to the "
            "2024-01-29 workbook; this downloader does not guess or repair unlinked URLs.",
            "The official archive exposes only 21 editions for 2024, with no linked "
            "editions between 2024-02-12 and 2024-09-30; missing weeks are not imputed.",
            "2025-10-06 through 2025-11-10 are edition labels, not actual catch-up "
            "publication dates; those rows become available only at project retrieval.",
        ],
        "methodology_breaks": [
            "FX tables first appear in the 2018-10-24 workbook despite the archive "
            "background describing the addition as of 2018-10-17.",
            "The announced December 2022 methodology improvement starts with the "
            "2022-12-12 edition in this frozen contract.",
            "The product schema changes separately on 2022-12-26: Exotics and Cross "
            "Currency merge into Other and must not be back-cast as separate categories.",
        ],
        "research_limitations": [
            "These are weekly SDR market-facing gross notional, ticket and dollar-volume "
            "aggregates, not signed spot-FX order flow.",
            "Currency rows have no buy/sell direction, no price, no forward points and no "
            "currency-specific participant-type split.",
            "NZD is included only inside an Other region in the official tables and cannot "
            "be mapped to an individual NZD factor.",
            "No row is a profitability result, alpha estimate or trading approval.",
        ],
        "sources": source_records,
    }
    manifest_payload = (json.dumps(manifest, indent=2) + "\n").encode()
    manifest_path = root / "cftc_weekly_swaps_manifest.json"
    _atomic_write(manifest_path, manifest_payload)
    _atomic_write(manifest_path.with_suffix(".sha256"), f"{_sha256(manifest_payload)}\n".encode())
    return manifest


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Download hash-audited CFTC Weekly Swaps FX activity editions"
    )
    parser.add_argument("--output", type=Path, default=Path("data/cftc_weekly_swaps"))
    parser.add_argument("--refresh", action="store_true")
    parser.add_argument("--timeout", type=float, default=45.0)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--delay", type=float, default=0.75)
    args = parser.parse_args(argv)
    if args.timeout <= 0 or not 0 <= args.retries <= 10 or args.delay < 0:
        parser.error("timeout must be positive, retries 0-10, and delay non-negative")
    try:
        manifest = download(
            args.output,
            refresh=args.refresh,
            timeout=args.timeout,
            retries=args.retries,
            delay=args.delay,
        )
    except (OSError, ValueError, httpx.HTTPError) as error:
        print(f"CFTC Weekly Swaps download failed: {error}", file=sys.stderr)
        return 1
    print(
        f"Saved {manifest['report_count']} official archive editions and "
        f"{manifest['normalized_rows']} exploratory FX-activity rows to {args.output}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
